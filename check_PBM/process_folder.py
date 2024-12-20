"""
입력받은 폴더 경로로부터 PBM을 검토하고, 오류 발생시 파일 리스트를 엑셀 파일을 생성
"""


import os
from natsort import natsorted
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm


def process_folder(input_path, output_path):
    """입력받은 폴더 경로로 PBM파일을 검토합니다"""
    error_pbm = []

    for root, _, files in os.walk(input_path):
        for file in tqdm(natsorted(files)):
            if not str(file).lower().endswith('.pbm'):
                continue
            pbm_path = os.path.join(root, file)

            with open(pbm_path, 'r', encoding='utf-8') as filename:
                soup = BeautifulSoup(filename, 'html.parser')
            level_tags = soup.find_all('level')

            for level in level_tags:
                content = level.get_text()

                if not '\n' in content:
                    continue
                error_pbm.append({
                    "file": file,
                    "path": pbm_path,
                    "text": content
                })

    if error_pbm:
        print("\nPBM에서 개행 발견")
        create_excel(error_pbm, output_path)
        print("\n엑셀파일 생성이 완료되었습니다\n")
    else:
        print("\nPBM에서 오류가 발생하지 않았습니다\n")


def create_excel(pbm_list, output_path):
    """입력받은 파일명을 엑셀로 내보냅니다"""
    wb = load_excel(output_path)
    ws = wb.active
    last_row = ws.max_row

    for idx, item in enumerate(pbm_list, 1):
        ws.cell(row=last_row + idx, column=1, value=idx)
        ws.cell(row=last_row + idx, column=2, value=item['file'])
        ws.cell(row=last_row + idx, column=3, value=item['path'])
        ws.cell(row=last_row + idx, column=4, value=item['text'])

    wb.save(output_path)


def load_excel(excel_file_path=str) -> Workbook:
    """엑셀 파일을 불러옵니다. 파일이 존재하지 않은 경우 HEADER를 추가하고 새 파일을 불러옵니다"""
    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active

        headers = ['연번', '파일명', '폴더경로', '내용']
        header_color = PatternFill(start_color='4f81bd',
                                   end_color='4f81bd', fill_type='solid')

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).fill = header_color
        wb.save(excel_file_path)
    wb = load_workbook(excel_file_path)

    return wb
